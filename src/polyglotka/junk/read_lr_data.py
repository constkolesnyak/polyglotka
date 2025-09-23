from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Any, Generator

import icecream
from funcy import lpluck_attr  # pyright: ignore
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from rich.console import Console
from rich.progress import BarColumn, Progress, SpinnerColumn, TextColumn


class LRRow(BaseModel):
    key: str = Field(validation_alias='Item key')
    type: str = Field(validation_alias='Item type')
    language: str = Field(validation_alias='Language')
    word: str = Field(validation_alias='Lemma')  # Not 'Word'
    part_of_speech: str | None = Field(validation_alias='Part of speech')
    source: str | None = Field(validation_alias='Source')
    video_title: str | None = Field(validation_alias='Video title')
    date: datetime = Field(validation_alias='Date created')

    @field_validator('date', mode='before')
    @classmethod
    def parse_date(cls, v: Any) -> datetime:
        return datetime.strptime(str(v), '%Y-%m-%d %H:%M')


def load_worksheet(excel_file: Path):
    worksheet = load_workbook(excel_file, read_only=True).active
    assert worksheet
    return worksheet


def read_lr_data(data_directory: str, progress: Progress) -> Generator[LRRow, None, None]:
    data_path = Path(data_directory)

    if not data_path.exists():
        raise FileNotFoundError(f'Directory not found: {data_directory}')
    excel_files = list(data_path.glob('lln_excel_items_*.xlsx'))
    if not excel_files:
        raise FileNotFoundError(f'No Excel files found in directory: {data_directory}')

    total_rows = 0
    for excel_file in excel_files:
        worksheet = load_worksheet(excel_file)
        total_rows += max(0, worksheet.max_row - 1)  # Subtract 1 for header row

    with progress:
        progress_task = progress.add_task('', total=total_rows)

        for excel_file in excel_files:
            worksheet = load_worksheet(excel_file)
            pluck_vals = partial(lpluck_attr, 'value')
            headers = pluck_vals(worksheet[1])

            for row_num in range(2, worksheet.max_row + 1):
                progress.update(progress_task, advance=1)
                yield LRRow(
                    **dict(zip(headers, pluck_vals(worksheet[row_num]), strict=True))
                )


def main() -> None:
    console = Console()
    progress = Progress(
        SpinnerColumn(style='bright_magenta'),
        BarColumn(complete_style='bright_magenta'),
        TextColumn('[bright_magenta]•'),
        TextColumn('[bright_magenta]{task.completed:,} / {task.total:,} words'),
        console=console,
        transient=True,
    )
    data_dir = "/Users/konst/google_drive/1. Inbox/polyglotka - inbox"

    total_items = 0
    for item in read_lr_data(data_dir, progress):
        total_items += 1
        if total_items <= 3:
            print(f'\nItem {total_items}:')
            icecream.ic(item.key)
            icecream.ic(item.type)
            icecream.ic(item.word)
            icecream.ic(item.language)
            icecream.ic(item.date)
        # else:
        #     exit()


if __name__ == "__main__":
    main()
